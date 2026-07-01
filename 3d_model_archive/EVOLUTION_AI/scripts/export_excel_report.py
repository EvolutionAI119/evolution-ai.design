"""
EVOLUTION AI - 25种车型参数对比测试报告 (Excel版)
"""

import os
import sys

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("安装openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl -q")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

from datetime import datetime

# 国标GB 1589-2016
GB = {"maxL": 5.3, "maxW": 2.55, "maxH": 4.0, "minGC": 0.13}

# 25种车型参数
MODELS = {
    "sedan_c": {"name":"C级轿车","brand":"蔚来ET7 / 奔驰E级 / 宝马5系","benchmark":"蔚来ET7 L=5101 WB=3060",
        "L":5.10,"W":1.99,"H":1.51,"WB":3.06,"FO":0.85,"RO":1.19,"GC":0.14,"WR":0.33,"TW":1.66,"AA":26,"CA":40,"WL":0.68,"WBulge":0.03},
    "sedan_d": {"name":"D级轿车","brand":"奔驰S级 / 宝马7系 / 奥迪A8L","benchmark":"奔驰S级 L=5290 WB=3216",
        "L":5.29,"W":1.92,"H":1.50,"WB":3.22,"FO":0.87,"RO":1.20,"GC":0.13,"WR":0.35,"TW":1.62,"AA":24,"CA":38,"WL":0.70,"WBulge":0.04},
    "compact": {"name":"紧凑轿车","brand":"奥迪A3 / 宝马i3 NK / 奔驰C级EV","benchmark":"奥迪A3L L=4606 WB=2730",
        "L":4.61,"W":1.81,"H":1.43,"WB":2.73,"FO":0.79,"RO":1.09,"GC":0.14,"WR":0.31,"TW":1.54,"AA":28,"CA":42,"WL":0.65,"WBulge":0.02},
    "wagon": {"name":"旅行车","brand":"蔚来ET5T / 奥迪A4 Avant / 奔驰C级旅行","benchmark":"蔚来ET5T L=4790 WB=2888",
        "L":4.79,"W":1.96,"H":1.50,"WB":2.89,"FO":0.80,"RO":1.10,"GC":0.14,"WR":0.33,"TW":1.64,"AA":26,"CA":55,"WL":0.68,"WBulge":0.02},
    "limo": {"name":"加长轿车","brand":"迈巴赫S级 / 奔驰S级 / 宝马7系长轴","benchmark":"迈巴赫S级 L=5470 WB=3396",
        "L":5.47,"W":1.92,"H":1.51,"WB":3.40,"FO":0.87,"RO":1.20,"GC":0.13,"WR":0.35,"TW":1.62,"AA":22,"CA":36,"WL":0.70,"WBulge":0.03},
    "suv_mid": {"name":"中型SUV","brand":"蔚来ES6 / 奔驰GLC EV / 宝马iX3 NK","benchmark":"蔚来ES6 L=4854 WB=2915",
        "L":4.85,"W":2.00,"H":1.70,"WB":2.92,"FO":0.81,"RO":1.12,"GC":0.20,"WR":0.37,"TW":1.66,"AA":30,"CA":45,"WL":0.82,"WBulge":0.02},
    "suv_full": {"name":"大型SUV","brand":"宝马X5 / 奔驰GLE / 奥迪Q7","benchmark":"宝马X5 L=5060 WB=3105",
        "L":5.06,"W":2.00,"H":1.78,"WB":3.11,"FO":0.82,"RO":1.13,"GC":0.22,"WR":0.40,"TW":1.68,"AA":32,"CA":48,"WL":0.90,"WBulge":0.01},
    "suv_compact": {"name":"紧凑SUV","brand":"奥迪Q3 / 奔驰GLA / 宝马X1","benchmark":"奥迪Q3 L=4498 WB=2680",
        "L":4.50,"W":1.85,"H":1.61,"WB":2.68,"FO":0.77,"RO":1.05,"GC":0.19,"WR":0.34,"TW":1.56,"AA":30,"CA":46,"WL":0.78,"WBulge":0.02},
    "suv_coupe": {"name":"轿跑SUV","brand":"保时捷Macan EV / 宝马X4 / 奔驰GLC Coupe","benchmark":"保时捷Macan EV L=4784 WB=2893",
        "L":4.78,"W":1.94,"H":1.62,"WB":2.89,"FO":0.79,"RO":1.10,"GC":0.19,"WR":0.37,"TW":1.62,"AA":26,"CA":38,"WL":0.74,"WBulge":0.04},
    "micro_suv": {"name":"小型SUV","brand":"奥迪Q2 / 宝马X2 / 奔驰GLB","benchmark":"奥迪Q2 L=4229 WB=2628",
        "L":4.23,"W":1.79,"H":1.55,"WB":2.63,"FO":0.67,"RO":0.93,"GC":0.18,"WR":0.31,"TW":1.50,"AA":32,"CA":50,"WL":0.74,"WBulge":0.01},
    "crossover": {"name":"跨界车","brand":"保时捷Macan / 奥迪Q3 Sportback / 宝马X2","benchmark":"保时捷Macan EV L=4784 WB=2893",
        "L":4.78,"W":1.94,"H":1.62,"WB":2.89,"FO":0.79,"RO":1.10,"GC":0.18,"WR":0.33,"TW":1.56,"AA":30,"CA":48,"WL":0.76,"WBulge":0.02},
    "coupe": {"name":"轿跑","brand":"保时捷911 / 奔驰AMG GT / 宝马8系","benchmark":"保时捷911 L=4519 WB=2450",
        "L":4.52,"W":1.85,"H":1.30,"WB":2.45,"FO":0.92,"RO":1.15,"GC":0.12,"WR":0.33,"TW":1.56,"AA":22,"CA":35,"WL":0.62,"WBulge":0.05},
    "sports": {"name":"超跑","brand":"法拉利296 / 保时捷918 / 奔驰AMG ONE","benchmark":"法拉利296 GTB L=4565 WB=2600",
        "L":4.57,"W":1.96,"H":1.19,"WB":2.60,"FO":0.87,"RO":1.10,"GC":0.10,"WR":0.34,"TW":1.64,"AA":20,"CA":32,"WL":0.55,"WBulge":0.06},
    "roadster": {"name":"敞篷跑车","brand":"保时捷718 / 奔驰SL / 宝马Z4","benchmark":"保时捷718 L=4379 WB=2475",
        "L":4.38,"W":1.80,"H":1.27,"WB":2.48,"FO":0.85,"RO":1.05,"GC":0.11,"WR":0.33,"TW":1.52,"AA":20,"CA":30,"WL":0.56,"WBulge":0.05},
    "hatchback": {"name":"掀背车","brand":"奥迪A3 Sportback / 奔驰A级 / 宝马1系","benchmark":"奥迪A3 SB L=4354 WB=2630",
        "L":4.35,"W":1.82,"H":1.46,"WB":2.63,"FO":0.72,"RO":1.00,"GC":0.14,"WR":0.31,"TW":1.54,"AA":28,"CA":50,"WL":0.66,"WBulge":0.02},
    "mpv": {"name":"MPV","brand":"奔驰V级 / 蔚来EL8 / 奥迪Q8","benchmark":"奔驰V级 L=5140 WB=3200",
        "L":5.14,"W":1.93,"H":1.88,"WB":3.20,"FO":0.82,"RO":1.12,"GC":0.16,"WR":0.34,"TW":1.62,"AA":24,"CA":52,"WL":0.85,"WBulge":0.01},
    "van": {"name":"轻客","brand":"奔驰Sprinter / 奔驰Vito / 奥迪e-tron GT","benchmark":"奔驰Sprinter L=5932 WB=3665",
        "L":5.93,"W":1.99,"H":1.98,"WB":3.67,"FO":0.95,"RO":1.31,"GC":0.18,"WR":0.35,"TW":1.72,"AA":22,"CA":58,"WL":0.95,"WBulge":0.01},
    "pickup": {"name":"皮卡","brand":"奔驰X级 / 奥迪RS Q8 / 保时捷Cayenne","benchmark":"奔驰X级 L=5340 WB=3150",
        "L":5.34,"W":1.92,"H":1.82,"WB":3.15,"FO":0.92,"RO":1.27,"GC":0.22,"WR":0.39,"TW":1.68,"AA":34,"CA":50,"WL":0.92,"WBulge":0.01},
    "ev_sedan": {"name":"电动轿车","brand":"蔚来ET5 / 奥迪e-tron GT / 保时捷Taycan","benchmark":"蔚来ET5 L=4790 WB=2888",
        "L":4.79,"W":1.96,"H":1.50,"WB":2.89,"FO":0.80,"RO":1.10,"GC":0.14,"WR":0.34,"TW":1.64,"AA":24,"CA":36,"WL":0.66,"WBulge":0.04},
    "ev_suv": {"name":"电动SUV","brand":"蔚来ES6 / 奔驰EQE SUV / 宝马iX3","benchmark":"蔚来ES6 L=4854 WB=2915",
        "L":4.85,"W":2.00,"H":1.70,"WB":2.92,"FO":0.81,"RO":1.12,"GC":0.18,"WR":0.36,"TW":1.66,"AA":28,"CA":44,"WL":0.80,"WBulge":0.02},
    "ev_compact": {"name":"电动紧凑","brand":"奥迪Q4 e-tron / 奔驰EQA / 宝马iX1","benchmark":"奥迪Q4 e-tron L=4588 WB=2765",
        "L":4.59,"W":1.87,"H":1.63,"WB":2.77,"FO":0.76,"RO":1.06,"GC":0.15,"WR":0.31,"TW":1.56,"AA":28,"CA":44,"WL":0.66,"WBulge":0.03},
    "ev_hatch": {"name":"电动两厢","brand":"保时捷Macan EV / 奥迪Q4 e-tron / 奔驰EQA","benchmark":"保时捷Macan EV L=4784 WB=2893",
        "L":4.78,"W":1.94,"H":1.62,"WB":2.89,"FO":0.79,"RO":1.10,"GC":0.16,"WR":0.30,"TW":1.52,"AA":30,"CA":50,"WL":0.70,"WBulge":0.02},
    "ev_pickup": {"name":"电动皮卡","brand":"奥迪Q8 e-tron / 奔驰EQG / 保时捷Cayenne E-Hybrid","benchmark":"奥迪Q8 e-tron L=4915 WB=2928",
        "L":4.92,"W":1.98,"H":1.68,"WB":2.93,"FO":0.84,"RO":1.15,"GC":0.22,"WR":0.40,"TW":1.72,"AA":32,"CA":48,"WL":0.94,"WBulge":0.01},
    "minicar": {"name":"微型车","brand":"MINI Cooper / smart fortwo / 奥迪AI:ME","benchmark":"MINI Cooper L=3858 WB=2520",
        "L":3.86,"W":1.76,"H":1.46,"WB":2.52,"FO":0.56,"RO":0.78,"GC":0.15,"WR":0.28,"TW":1.48,"AA":32,"CA":55,"WL":0.72,"WBulge":0.01},
    "kei_car": {"name":"K-Car","brand":"smart fortwo / MINI / 奥迪AI:ME","benchmark":"smart fortwo L=2695 WB=1873",
        "L":2.70,"W":1.66,"H":1.56,"WB":1.87,"FO":0.35,"RO":0.48,"GC":0.15,"WR":0.27,"TW":1.40,"AA":34,"CA":58,"WL":0.78,"WBulge":0.01}
}

# 真实市场数据(mm)
REAL_DATA = {
    "sedan_c":{"name":"蔚来ET7","L":5101,"W":1987,"H":1509,"WB":3060},
    "sedan_d":{"name":"奔驰S级","L":5290,"W":1921,"H":1503,"WB":3216},
    "compact":{"name":"奥迪A3L","L":4606,"W":1814,"H":1432,"WB":2730},
    "wagon":{"name":"蔚来ET5T","L":4790,"W":1960,"H":1499,"WB":2888},
    "limo":{"name":"迈巴赫S级","L":5470,"W":1921,"H":1510,"WB":3396},
    "suv_mid":{"name":"蔚来ES6","L":4854,"W":1995,"H":1703,"WB":2915},
    "suv_full":{"name":"宝马X5","L":5060,"W":2004,"H":1779,"WB":3105},
    "suv_compact":{"name":"奥迪Q3","L":4498,"W":1848,"H":1614,"WB":2680},
    "suv_coupe":{"name":"Macan EV","L":4784,"W":1938,"H":1622,"WB":2893},
    "micro_suv":{"name":"奥迪Q2","L":4229,"W":1785,"H":1548,"WB":2628},
    "crossover":{"name":"Macan EV","L":4784,"W":1938,"H":1622,"WB":2893},
    "coupe":{"name":"保时捷911","L":4519,"W":1852,"H":1298,"WB":2450},
    "sports":{"name":"法拉利296","L":4565,"W":1958,"H":1186,"WB":2600},
    "roadster":{"name":"保时捷718","L":4379,"W":1801,"H":1272,"WB":2475},
    "hatchback":{"name":"奥迪A3 SB","L":4354,"W":1815,"H":1458,"WB":2630},
    "mpv":{"name":"奔驰V级","L":5140,"W":1928,"H":1880,"WB":3200},
    "van":{"name":"Sprinter","L":5932,"W":1993,"H":1980,"WB":3665},
    "pickup":{"name":"奔驰X级","L":5340,"W":1920,"H":1819,"WB":3150},
    "ev_sedan":{"name":"蔚来ET5","L":4790,"W":1960,"H":1499,"WB":2888},
    "ev_suv":{"name":"蔚来ES6","L":4854,"W":1995,"H":1703,"WB":2915},
    "ev_compact":{"name":"Q4 e-tron","L":4588,"W":1865,"H":1626,"WB":2765},
    "ev_hatch":{"name":"Macan EV","L":4784,"W":1938,"H":1622,"WB":2893},
    "ev_pickup":{"name":"Q8 e-tron","L":4915,"W":1976,"H":1680,"WB":2928},
    "minicar":{"name":"MINI Cooper","L":3858,"W":1756,"H":1460,"WB":2520},
    "kei_car":{"name":"smart fortwo","L":2695,"W":1663,"H":1555,"WB":1873}
}

# 样式定义
HEADER_FONT = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
DATA_FONT = Font(name="微软雅黑", size=9)
PASS_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
WARN_FILL = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
TITLE_FONT = Font(name="微软雅黑", size=14, bold=True, color="1A237E")
SUBTITLE_FONT = Font(name="微软雅黑", size=10, color="5A6E8A")
THIN_BORDER = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0')
)
CENTER = Alignment(horizontal='center', vertical='center')

PARAM_KEYS = ["L","W","H","WB","FO","RO","GC","WR","TW","AA","CA","WL","WBulge"]
PARAM_NAMES = {"L":"车长(m)","W":"车宽(m)","H":"车高(m)","WB":"轴距(m)","FO":"前悬(m)","RO":"后悬(m)",
               "GC":"离地间隙(m)","WR":"轮径(m)","TW":"轮距(m)","AA":"A柱角(°)","CA":"C柱角(°)",
               "WL":"腰线高(m)","WBulge":"轮拱凸出(m)"}

CATEGORIES = {
    "轿车":["sedan_c","sedan_d","compact","wagon","limo"],
    "SUV":["suv_mid","suv_full","suv_compact","suv_coupe","micro_suv","crossover"],
    "跑车":["coupe","sports","roadster","hatchback"],
    "MPV/皮卡":["mpv","van","pickup"],
    "电动":["ev_sedan","ev_suv","ev_compact","ev_hatch","ev_pickup"],
    "微型":["minicar","kei_car"]
}


def style_header(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def style_data(ws, row, cols, fill=None):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = DATA_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill


def generate_excel():
    wb = openpyxl.Workbook()

    # ========== Sheet 1: 参数总表 ==========
    ws1 = wb.active
    ws1.title = "参数总表"
    ws1.merge_cells('A1:P1')
    ws1['A1'] = "EVOLUTION AI - 25种车型参数总表"
    ws1['A1'].font = TITLE_FONT
    ws1.merge_cells('A2:P2')
    ws1['A2'] = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | 国标: GB 1589-2016"
    ws1['A2'].font = SUBTITLE_FONT

    headers = ["序号","车型ID","车型名称","标杆品牌","车长(m)","车宽(m)","车高(m)","轴距(m)",
               "前悬(m)","后悬(m)","离地间隙(m)","轮径(m)","轮距(m)","A柱角(°)","C柱角(°)",
               "腰线高(m)","轮拱凸出(m)"]
    for i, h in enumerate(headers, 1):
        ws1.cell(row=4, column=i, value=h)
    style_header(ws1, 4, len(headers))

    row = 5
    for idx, (key, m) in enumerate(MODELS.items(), 1):
        ws1.cell(row=row, column=1, value=idx)
        ws1.cell(row=row, column=2, value=key)
        ws1.cell(row=row, column=3, value=m["name"])
        ws1.cell(row=row, column=4, value=m["brand"])
        for j, pk in enumerate(PARAM_KEYS):
            ws1.cell(row=row, column=5+j, value=m[pk])
        style_data(ws1, row, len(headers))
        row += 1

    for c in range(1, len(headers)+1):
        ws1.column_dimensions[get_column_letter(c)].width = 14

    # ========== Sheet 2: 国标合规检查 ==========
    ws2 = wb.create_sheet("国标合规检查")
    ws2.merge_cells('A1:G1')
    ws2['A1'] = "国标GB 1589-2016合规性检查"
    ws2['A1'].font = TITLE_FONT

    headers2 = ["车型","车长(m)","车宽(m)","车高(m)","离地间隙(m)","合规状态","违规项"]
    for i, h in enumerate(headers2, 1):
        ws2.cell(row=3, column=i, value=h)
    style_header(ws2, 3, len(headers2))

    row = 4
    pass_count = 0
    for key, m in MODELS.items():
        issues = []
        if m["L"] > GB["maxL"]: issues.append(f"车长{m['L']:.2f}m>{GB['maxL']}m")
        if m["W"] > GB["maxW"]: issues.append(f"车宽{m['W']:.2f}m>{GB['maxW']}m")
        if m["H"] > GB["maxH"]: issues.append(f"车高{m['H']:.2f}m>{GB['maxH']}m")
        if m["GC"] < GB["minGC"]: issues.append(f"离地{m['GC']:.2f}m<{GB['minGC']}m")

        status = "合规" if not issues else "违规"
        fill = PASS_FILL if not issues else FAIL_FILL
        if not issues: pass_count += 1

        ws2.cell(row=row, column=1, value=m["name"])
        ws2.cell(row=row, column=2, value=m["L"])
        ws2.cell(row=row, column=3, value=m["W"])
        ws2.cell(row=row, column=4, value=m["H"])
        ws2.cell(row=row, column=5, value=m["GC"])
        ws2.cell(row=row, column=6, value=status)
        ws2.cell(row=row, column=7, value="; ".join(issues) if issues else "-")
        style_data(ws2, row, len(headers2), fill)
        row += 1

    ws2.cell(row=row+1, column=1, value="合规率")
    ws2.cell(row=row+1, column=2, value=f"{pass_count}/{len(MODELS)} ({pass_count/len(MODELS)*100:.0f}%)")
    ws2.cell(row=row+1, column=2).font = Font(name="微软雅黑", size=11, bold=True, color="1A237E")

    for c in range(1, len(headers2)+1):
        ws2.column_dimensions[get_column_letter(c)].width = 16

    # ========== Sheet 3: 几何一致性 ==========
    ws3 = wb.create_sheet("几何一致性")
    ws3.merge_cells('A1:F1')
    ws3['A1'] = "几何一致性检查 (FO+WB+RO≈L)"
    ws3['A1'].font = TITLE_FONT

    headers3 = ["车型","FO(m)","WB(m)","RO(m)","FO+WB+RO","L(m)","差值(m)","状态"]
    for i, h in enumerate(headers3, 1):
        ws3.cell(row=3, column=i, value=h)
    style_header(ws3, 3, len(headers3))

    row = 4
    for key, m in MODELS.items():
        total = m["FO"] + m["WB"] + m["RO"]
        diff = total - m["L"]
        status = "通过" if abs(diff) < 0.15 else "异常"
        fill = PASS_FILL if abs(diff) < 0.15 else FAIL_FILL
        ws3.cell(row=row, column=1, value=m["name"])
        ws3.cell(row=row, column=2, value=m["FO"])
        ws3.cell(row=row, column=3, value=m["WB"])
        ws3.cell(row=row, column=4, value=m["RO"])
        ws3.cell(row=row, column=5, value=round(total, 3))
        ws3.cell(row=row, column=6, value=m["L"])
        ws3.cell(row=row, column=7, value=round(diff, 3))
        ws3.cell(row=row, column=8, value=status)
        style_data(ws3, row, len(headers3), fill)
        row += 1

    for c in range(1, len(headers3)+1):
        ws3.column_dimensions[get_column_letter(c)].width = 14

    # ========== Sheet 4: 比例分析 ==========
    ws4 = wb.create_sheet("比例分析")
    ws4.merge_cells('A1:G1')
    ws4['A1'] = "比例合理性分析"
    ws4['A1'].font = TITLE_FONT

    headers4 = ["车型","轴距/车长","前悬占比","高宽比","轮心高(m)","腰线(m)","轮距/车宽","状态"]
    for i, h in enumerate(headers4, 1):
        ws4.cell(row=3, column=i, value=h)
    style_header(ws4, 3, len(headers4))

    row = 4
    for key, m in MODELS.items():
        wb_ratio = m["WB"] / m["L"]
        fo_ratio = m["FO"] / (m["FO"] + m["RO"])
        hw_ratio = m["H"] / m["W"]
        wc = m["GC"] + m["WR"]
        tw_ratio = m["TW"] / m["W"]

        issues = []
        if wb_ratio < 0.55 or wb_ratio > 0.70: issues.append("轴距比异常")
        if fo_ratio < 0.35 or fo_ratio > 0.55: issues.append("前悬比异常")
        if hw_ratio < 0.60 or hw_ratio > 1.0: issues.append("高宽比异常")
        status = "正常" if not issues else "; ".join(issues)
        fill = PASS_FILL if not issues else WARN_FILL

        ws4.cell(row=row, column=1, value=m["name"])
        ws4.cell(row=row, column=2, value=round(wb_ratio, 3))
        ws4.cell(row=row, column=3, value=round(fo_ratio, 3))
        ws4.cell(row=row, column=4, value=round(hw_ratio, 3))
        ws4.cell(row=row, column=5, value=round(wc, 3))
        ws4.cell(row=row, column=6, value=m["WL"])
        ws4.cell(row=row, column=7, value=round(tw_ratio, 3))
        ws4.cell(row=row, column=8, value=status)
        style_data(ws4, row, len(headers4), fill)
        row += 1

    for c in range(1, len(headers4)+1):
        ws4.column_dimensions[get_column_letter(c)].width = 14

    # ========== Sheet 5: 标杆验证 ==========
    ws5 = wb.create_sheet("标杆数据验证")
    ws5.merge_cells('A1:I1')
    ws5['A1'] = "标杆车型数据验证 (模型参数 vs 真实市场数据)"
    ws5['A1'].font = TITLE_FONT

    headers5 = ["车型","标杆车型","模型L(mm)","真实L(mm)","L差(mm)","模型W(mm)","真实W(mm)","W差(mm)",
                "模型H(mm)","真实H(mm)","H差(mm)","模型WB(mm)","真实WB(mm)","WB差(mm)","最大偏差(mm)","状态"]
    for i, h in enumerate(headers5, 1):
        ws5.cell(row=3, column=i, value=h)
    style_header(ws5, 3, len(headers5))

    row = 4
    for key, m in MODELS.items():
        rd = REAL_DATA.get(key)
        if not rd:
            ws5.cell(row=row, column=1, value=m["name"])
            ws5.cell(row=row, column=2, value="-")
            ws5.cell(row=row, column=16, value="无标杆数据")
            style_data(ws5, row, len(headers5), WARN_FILL)
            row += 1
            continue

        ml = m["L"]*1000; rl = rd["L"]; dl = abs(ml-rl)
        mw = m["W"]*1000; rw = rd["W"]; dw = abs(mw-rw)
        mh = m["H"]*1000; rh = rd["H"]; dh = abs(mh-rh)
        mwb = m["WB"]*1000; rwb = rd["WB"]; dwb = abs(mwb-rwb)
        max_err = max(dl, dw, dh, dwb)
        status = "精确" if max_err < 10 else "合格" if max_err < 50 else "偏差"
        fill = PASS_FILL if max_err < 50 else WARN_FILL

        ws5.cell(row=row, column=1, value=m["name"])
        ws5.cell(row=row, column=2, value=rd["name"])
        ws5.cell(row=row, column=3, value=int(ml))
        ws5.cell(row=row, column=4, value=rl)
        ws5.cell(row=row, column=5, value=int(dl))
        ws5.cell(row=row, column=6, value=int(mw))
        ws5.cell(row=row, column=7, value=rw)
        ws5.cell(row=row, column=8, value=int(dw))
        ws5.cell(row=row, column=9, value=int(mh))
        ws5.cell(row=row, column=10, value=rh)
        ws5.cell(row=row, column=11, value=int(dh))
        ws5.cell(row=row, column=12, value=int(mwb))
        ws5.cell(row=row, column=13, value=rwb)
        ws5.cell(row=row, column=14, value=int(dwb))
        ws5.cell(row=row, column=15, value=int(max_err))
        ws5.cell(row=row, column=16, value=status)
        style_data(ws5, row, len(headers5), fill)
        row += 1

    for c in range(1, len(headers5)+1):
        ws5.column_dimensions[get_column_letter(c)].width = 14

    # ========== Sheet 6: 同类对比 ==========
    ws6 = wb.create_sheet("同类车型对比")
    ws6.merge_cells('A1:H1')
    ws6['A1'] = "同类车型参数对比"
    ws6['A1'].font = TITLE_FONT

    row = 3
    for cat_name, cat_keys in CATEGORIES.items():
        ws6.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        ws6.cell(row=row, column=1, value=f"【{cat_name}】")
        ws6.cell(row=row, column=1).font = Font(name="微软雅黑", size=11, bold=True, color="1A237E")
        row += 1

        headers6 = ["车型","车长(m)","车宽(m)","车高(m)","轴距(m)","离地间隙(m)","L偏差","备注"]
        for i, h in enumerate(headers6, 1):
            ws6.cell(row=row, column=i, value=h)
        style_header(ws6, row, len(headers6))
        row += 1

        cat_models = [(k, MODELS[k]) for k in cat_keys if k in MODELS]
        avg_l = sum(m["L"] for _, m in cat_models) / len(cat_models)

        for key, m in cat_models:
            diff_l = m["L"] - avg_l
            ws6.cell(row=row, column=1, value=m["name"])
            ws6.cell(row=row, column=2, value=m["L"])
            ws6.cell(row=row, column=3, value=m["W"])
            ws6.cell(row=row, column=4, value=m["H"])
            ws6.cell(row=row, column=5, value=m["WB"])
            ws6.cell(row=row, column=6, value=m["GC"])
            ws6.cell(row=row, column=7, value=round(diff_l, 2))
            ws6.cell(row=row, column=8, value=m["brand"])
            style_data(ws6, row, len(headers6))
            row += 1

        # 平均行
        ws6.cell(row=row, column=1, value="平均")
        ws6.cell(row=row, column=1).font = Font(name="微软雅黑", size=9, bold=True)
        avg_w = sum(m["W"] for _, m in cat_models) / len(cat_models)
        avg_h = sum(m["H"] for _, m in cat_models) / len(cat_models)
        avg_wb = sum(m["WB"] for _, m in cat_models) / len(cat_models)
        avg_gc = sum(m["GC"] for _, m in cat_models) / len(cat_models)
        ws6.cell(row=row, column=2, value=round(avg_l, 2))
        ws6.cell(row=row, column=3, value=round(avg_w, 2))
        ws6.cell(row=row, column=4, value=round(avg_h, 2))
        ws6.cell(row=row, column=5, value=round(avg_wb, 2))
        ws6.cell(row=row, column=6, value=round(avg_gc, 2))
        style_data(ws6, row, len(headers6), PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"))
        row += 2

    for c in range(1, 9):
        ws6.column_dimensions[get_column_letter(c)].width = 14

    # ========== Sheet 7: 尺寸排名 ==========
    ws7 = wb.create_sheet("尺寸排名")
    ws7.merge_cells('A1:D1')
    ws7['A1'] = "尺寸排名"
    ws7['A1'].font = TITLE_FONT

    row = 3
    for param, name in [("L","车长"),("W","车宽"),("H","车高"),("WB","轴距"),("GC","离地间隙")]:
        ws7.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws7.cell(row=row, column=1, value=f"{name}排名")
        ws7.cell(row=row, column=1).font = Font(name="微软雅黑", size=11, bold=True, color="1A237E")
        row += 1

        for i, h in enumerate(["排名","车型",f"{name}(m)",f"{name}(mm)"], 1):
            ws7.cell(row=row, column=i, value=h)
        style_header(ws7, row, 4)
        row += 1

        sorted_models = sorted(MODELS.items(), key=lambda x: x[1][param], reverse=True)
        for rank, (key, m) in enumerate(sorted_models, 1):
            ws7.cell(row=row, column=1, value=rank)
            ws7.cell(row=row, column=2, value=m["name"])
            ws7.cell(row=row, column=3, value=m[param])
            ws7.cell(row=row, column=4, value=int(m[param]*1000))
            style_data(ws7, row, 4)
            row += 1
        row += 1

    for c in range(1, 5):
        ws7.column_dimensions[get_column_letter(c)].width = 16

    # ========== 保存 ==========
    output_dir = os.path.join(os.path.dirname(__file__), "output")
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, "EVOLUTION_AI_车型参数对比报告.xlsx")
    wb.save(output_path)
    print(f"Excel报告已保存到: {output_path}")
    return output_path


if __name__ == "__main__":
    path = generate_excel()
    print(f"文件大小: {os.path.getsize(path)/1024:.1f} KB")
